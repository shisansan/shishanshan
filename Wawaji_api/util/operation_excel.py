#!/usr/bin/env python
import xlrd
from xlutils.copy import copy
import openpyxl

class OperationExcel:

    def __init__(self, filename=None, sheet_name=None):
        if filename:
            self.filename = filename
            # self.sheet_id = sheet_id
            self.sheet_name = sheet_name
        else:
            self.filename = '../dataconfig/interface.xlsx'
            self.sheet_name = 'Sheet1'
            # self.sheet_id = 0
        self.data = self.get_data()

    #获取sheets的内容
    def get_data(self):
        data = xlrd.open_workbook(self.filename)
        # tables = data.sheets()[self.sheet_id]
        tables = data.sheet_by_name(self.sheet_name)
        return tables

    # def get_data(self):
    #     wb = openpyxl.load_workbook(self.filename)
    #     ws = wb[self.sheet_name]
    #     return ws

    #获取sheets的行数
    def get_line(self):
        tables = self.data
        return tables.nrows

    #获取某一个单元格
    def get_cell_value(self, row, col):
        return self.data.cell_value(row, col)

    # #写入数据
    # def write_value(self, row, col, value):
    #     read_data = xlrd.open_workbook(self.filename)
    #     write_data = copy(read_data)
    #     sheet_data = write_data.get_sheet(self.sheet_id)
    #     sheet_data.write(row, col, value)
    #     write_data.save(self.filename)

    #写入数据
    def write_value(self, row, col, value):
        wb = openpyxl.load_workbook(self.filename)
        ws = wb[self.sheet_name] #激活sheet
        ws.cell(row, col).value = value
        wb.save(self.filename)

    #解决数据依赖之：根据对应的case_id找到对应行的内容
    def get_rows_data(self, case_id):
        row_num = self.get_row_num(case_id)
        rows_data = self.get_row_values(row_num)
        return rows_data

    #解决数据依赖之：根据对应的case_id找到对应的行号
    def get_row_num(self, case_id):
        num = 0
        cols_data = self.get_cols_data()
        for col_data in cols_data:
            if case_id in col_data:
                return num
        num += 1

    #解决数据依赖之：根据行号，找到该行的内容
    def get_row_values(self, row):
        tables = self.data
        row_data = tables.row_values(row)
        return row_data


    #获取某一列的内容
    def get_cols_data(self, col_id=None):
        if col_id != None:
            cols = self.data.col_values(col_id)
        else:
            cols = self.data.col_values(0)
        return cols



if __name__ == '__main__':
    opers = OperationExcel()
    print(opers.get_data().nrows)
    print(opers.get_line())
    print(opers.get_cell_value(1, 1))
    opers.write_value(2, 13, 'qqq')